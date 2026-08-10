from sqlalchemy.orm import Session
from app.models import Product, Category
from typing import Optional, List, Tuple
import io
import pandas as pd
import unicodedata
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

class ProductImportService:
    """Serviço de Importação/Exportação de Produtos"""
    
    @staticmethod
    def _normalizar_texto(value) -> str:
        if pd.isna(value):
            return ""

        text = str(value).strip().lower()
        text = unicodedata.normalize("NFKD", text)
        return "".join(char for char in text if not unicodedata.combining(char))
    
    @staticmethod
    def _parse_preco(value) -> float:
        if pd.isna(value):
            raise ValueError("Preco vazio")

        if isinstance(value, (int, float)):
            return float(value)

        text = str(value).strip().replace("R$", "").replace(" ", "")

        if "," in text and "." in text:
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", ".")

        return float(text)
    
    @staticmethod
    def _is_empty(value) -> bool:
        return pd.isna(value) or str(value).strip() == ""
    
    @staticmethod
    def _is_category_row(row: pd.Series) -> bool:
        non_empty = [
            value for value in row.tolist()
            if not ProductImportService._is_empty(value)
        ]

        if len(non_empty) != 1:
            return False

        text = str(non_empty[0]).strip()
        normalized = ProductImportService._normalizar_texto(text)
        headers = {"cod", "codigo", "cod.", "nome", "valor", "aliquota", "tipo"}

        return normalized not in headers and not text.replace(".", "", 1).isdigit()
    
    @staticmethod
    def _find_header_columns(row: pd.Series) -> Optional[dict]:
        columns = {}

        for idx, value in row.items():
            text = ProductImportService._normalizar_texto(value)

            if text in {"cod", "codigo", "cod."}:
                columns["codigo"] = idx
            elif text == "nome":
                columns["nome"] = idx
            elif text == "valor":
                columns["valor"] = idx
            elif text in {"aliquota", "aliq"}:
                columns["aliquota"] = idx
            elif text == "tipo":
                columns["tipo"] = idx

        if "nome" in columns and "valor" in columns:
            return columns

        return None
    
    @staticmethod
    def _extract_price_from_layout(row: pd.Series, columns: dict) -> float:
        valor_idx = columns["valor"]
        end_idx = columns.get("aliquota") or len(row)

        for idx in range(valor_idx, end_idx):
            value = row.iloc[idx]

            if ProductImportService._is_empty(value):
                continue

            normalized = ProductImportService._normalizar_texto(value)
            if normalized in {"r$", "rs"}:
                continue

            return ProductImportService._parse_preco(value)

        raise ValueError("Preco vazio")
    
    @staticmethod
    def normalizar_excel(contents: bytes) -> pd.DataFrame:
        """Aceita o modelo simples e o layout por categorias do sistema atual."""

        simple_df = pd.read_excel(io.BytesIO(contents))
        simple_columns = {
            ProductImportService._normalizar_texto(column): column
            for column in simple_df.columns
        }

        if "nome" in simple_columns and (
            "preco" in simple_columns or "valor" in simple_columns
        ):
            nome_col = simple_columns["nome"]
            preco_col = simple_columns.get("preco") or simple_columns["valor"]
            categoria_col = simple_columns.get("categoria")

            data = []
            for _, row in simple_df.iterrows():
                if ProductImportService._is_empty(row.get(nome_col)):
                    continue

                data.append({
                    "nome": row.get(nome_col),
                    "preco": row.get(preco_col),
                    "categoria": row.get(categoria_col) if categoria_col else None,
                })

            return pd.DataFrame(data)

        raw_df = pd.read_excel(io.BytesIO(contents), header=None)
        current_category = None
        header_columns = None
        data = []

        for idx, row in raw_df.iterrows():
            if row.dropna().empty:
                continue

            detected_header = ProductImportService._find_header_columns(row)
            if detected_header:
                header_columns = detected_header
                continue

            if ProductImportService._is_category_row(row):
                current_category = str(row.dropna().iloc[0]).strip()
                continue

            if not header_columns:
                continue

            nome = row.iloc[header_columns["nome"]]
            if ProductImportService._is_empty(nome):
                continue

            try:
                preco = ProductImportService._extract_price_from_layout(
                    row,
                    header_columns,
                )
            except ValueError:
                preco = None

            data.append({
                "nome": str(nome).strip(),
                "preco": preco,
                "categoria": current_category,
                "_linha_excel": idx + 1,
            })

        return pd.DataFrame(data)
    
    @staticmethod
    def validar_dados(df: pd.DataFrame) -> Tuple[bool, List[str]]:
        """Valida estrutura do DataFrame"""
        
        erros = []
        
        # Verifica colunas obrigatórias
        colunas_obrigatorias = ['nome', 'preco']
        for col in colunas_obrigatorias:
            if col not in df.columns:
                erros.append(f"Coluna obrigatória ausente: {col}")
        
        if erros:
            return False, erros
        
        # Valida dados
        for idx, row in df.iterrows():
            linha = int(row.get('_linha_excel', idx + 2))
            
            # Nome vazio
            if pd.isna(row['nome']) or str(row['nome']).strip() == '':
                erros.append(f"Linha {linha}: Nome vazio")
            
            # Preço inválido
            try:
                preco = ProductImportService._parse_preco(row['preco'])
                if preco <= 0:
                    erros.append(f"Linha {linha}: Preço deve ser > 0")
            except (ValueError, TypeError):
                erros.append(f"Linha {linha}: Preço inválido")
        
        return len(erros) == 0, erros
    
    @staticmethod
    def importar_produtos(
        db: Session,
        company_id: int,
        df: pd.DataFrame
    ) -> dict:
        """
        Importa produtos do DataFrame
        
        Retorna:
            {
                "importados": int,
                "atualizados": int,
                "erros": list,
                "total": int
            }
        """
        
        importados = 0
        atualizados = 0
        erros = []
        
        for idx, row in df.iterrows():
            try:
                linha = int(row.get('_linha_excel', idx + 2))
                nome = str(row['nome']).strip()
                preco = ProductImportService._parse_preco(row['preco'])
                categoria_nome = row.get('categoria', None)
                
                # Validações básicas
                if not nome:
                    erros.append(f"Linha {linha}: Nome vazio")
                    continue
                
                if preco <= 0:
                    erros.append(f"Linha {linha}: Preço inválido")
                    continue
                
                # Busca ou cria categoria
                categoria_id = None
                if categoria_nome and pd.notna(categoria_nome):
                    categoria_nome = str(categoria_nome).strip()
                    categoria = db.query(Category).filter(
                        Category.nome == categoria_nome,
                        Category.company_id == company_id
                    ).first()
                    
                    if not categoria:
                        categoria = Category(nome=categoria_nome, company_id=company_id)
                        db.add(categoria)
                        db.flush()
                    
                    categoria_id = categoria.id
                
                # Busca produto existente
                produto = db.query(Product).filter(
                    Product.nome == nome,
                    Product.company_id == company_id
                ).first()
                
                if produto:
                    # Atualiza
                    produto.preco = preco
                    if categoria_id:
                        produto.categoria_id = categoria_id
                    db.add(produto)
                    atualizados += 1
                else:
                    # Cria novo
                    produto = Product(
                        nome=nome,
                        preco=preco,
                        categoria_id=categoria_id,
                        company_id=company_id
                    )
                    db.add(produto)
                    importados += 1
            
            except Exception as e:
                erros.append(f"Linha {idx + 2}: {str(e)}")
        
        db.commit()
        
        return {
            "importados": importados,
            "atualizados": atualizados,
            "erros": erros,
            "total": len(erros) > 0 and 0 or (importados + atualizados)
        }
    
    @staticmethod
    def gerar_template_excel() -> bytes:
        """Gera arquivo Excel modelo"""
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Produtos"
        
        # Header
        headers = ['nome', 'categoria', 'preco']
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        # Exemplo
        ws.cell(row=2, column=1).value = "Produto A"
        ws.cell(row=2, column=2).value = "Bebidas"
        ws.cell(row=2, column=3).value = 10.50
        
        # Ajusta largura
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 12
        
        # Salva em BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    @staticmethod
    def exportar_produtos(
        db: Session,
        company_id: int
    ) -> bytes:
        """Exporta todos os produtos para Excel"""
        
        products = db.query(Product).filter(
            Product.company_id == company_id
        ).all()
        
        # Cria DataFrame
        data = []
        for product in products:
            categoria = product.category.nome if product.category else ""
            data.append({
                'nome': product.nome,
                'categoria': categoria,
                'preco': product.preco
            })
        
        df = pd.DataFrame(data)
        
        # Cria Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Produtos"
        
        # Header
        headers = ['nome', 'categoria', 'preco']
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        # Dados
        for row_idx, row in df.iterrows():
            ws.cell(row=row_idx + 2, column=1).value = row['nome']
            ws.cell(row=row_idx + 2, column=2).value = row['categoria']
            ws.cell(row=row_idx + 2, column=3).value = row['preco']
        
        # Ajusta largura
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 12
        
        # Salva em BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
